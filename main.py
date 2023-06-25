import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
from openpyxl.styles import PatternFill
import pandas as pd

colB_list = []
colC_list = []
colF_list = []
colG_list = []
dur_list = []

# Load the workbook
workbook = openpyxl.load_workbook('File.xlsx')
worksheet = workbook.active

# Set the yellow fill style
lightBlue_fill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')

# Determine the column to insert before
column_letter = openpyxl.utils.get_column_letter(1)

# Insert the new column
for x in range(2, 9):
    worksheet.insert_cols(x)

# Iterate through the rows in the worksheet
for row in range(1, worksheet.max_row + 1):

    value = worksheet.cell(row=row, column=1).value
    if value is None:
        continue
    parts = value.split(" ", 1)
    worksheet.cell(row=row, column=1).value = parts[0]
    worksheet.cell(row=row, column=2).value = parts[1] if len(parts) > 1 else ""

# Iterate through the rows in the worksheet
for row in range(1, worksheet.max_row + 1):
    value = worksheet.cell(row=row, column=2).value
    if value is None:
        continue
    parts = value.split("-", 1)
    worksheet.cell(row=row, column=2).value = parts[0]
    worksheet.cell(row=row, column=3).value = parts[1] if len(parts) > 1 else ""

# Set the starting row
row = 2

# Get the first date
date = datetime.strptime(worksheet.cell(row, 1).value, '%Y-%m-%d').date()

# Iterate over all the rows in the sheet
while worksheet.cell(row, 1).value:

    # Get the date from the current row
    current_date = datetime.strptime(worksheet.cell(row, 1).value, '%Y-%m-%d').date()

    # If the date has changed, insert a new row
    if current_date != date:
        worksheet.insert_rows(row)
        for col in range(1, worksheet.max_column + 1):
            worksheet.cell(row, col).value = ""
        worksheet.cell(row, 1).value = "."

        # Loop through all cells in the row
        #for cell in worksheet[row]:
            # Set the fill color to yellow
            # cell.fill = openpyxl.styles.PatternFill(start_color="000000", end_color="000000", fill_type="solid")
        date = current_date

    # Move to the next row
    row += 1


def rowColors():
    # Loop through all the rows in the worksheet
    for row in worksheet.iter_rows(min_row=1, max_col=1):
        # Check the value of the cell in column A
        if row[0].value == ".":
            # If the value is None, the cell is blank
            # Access cell A1
            cellcount = row[0].row
            previous = cellcount - 1
            next = cellcount + 1
            currentcell = f"G{previous}"
            nextcell = f"H{next}"

            # print(currentcell)
            cell2 = worksheet[currentcell]
            cell3 = worksheet[nextcell]
            mostrecent = worksheet["F2"]

            # Create a font object and set its color to white
            font = openpyxl.styles.Font(color="FFFFFF")

            # Get the last row index for the worksheet
            last_row = worksheet[f"B{worksheet.max_row - 1}"]

            # Set the fill color to red
            fill = PatternFill(start_color="FF0000", end_color="FF0000", patternType="solid")
            cell2.fill = fill
            cell2.font = font

            cell3.fill = fill
            cell3.font = font

            mostrecent.fill = fill
            mostrecent.font = font

            last_row.fill = fill
            last_row.font = font


# rowColors()

# Get the column dimensions
column_dim = worksheet.column_dimensions["B"]

# Set the width of the column to fit the text
column_dim.width = max(len(str(cell.value)) for cell in worksheet["B"])


def round_column_b():
    # Iterate through the rows of the worksheet
    for row in worksheet.iter_rows(min_row=2, max_col=2, max_row=worksheet.max_row, values_only=True):
        # Get the cell value from column B
        cell_value = row[1]

        if cell_value == "time":
            pass

        if row[0] == "Flight":
            pass

        if row[0] == ".":
            colB_list.append(".")

        else:
            new_date = f"{row[0]} {cell_value}"

            if new_date == "None None":
                pass
            else:
                # print(new_date)
                # Print the cell value
                df = pd.DataFrame(
                    columns=["datetime"],
                    data=pd.date_range(new_date, periods=1, freq="s"))

                df["minute_datetime"] = df["datetime"].dt.round("min")

                time = df["minute_datetime"].dt.strftime("%H:%M")

                contime = time.to_string()
                round_time = contime.split("   ")[1]
                colB_list.append(round_time.split(" ")[1])
    print(colB_list)

    for row, value in zip(worksheet.iter_rows(min_col=4, max_col=4, min_row=2), colB_list):
        row[0].value = value


def round_column_c():
    # Iterate through the rows of the worksheet
    for row in worksheet.iter_rows(min_row=2, max_col=4, max_row=worksheet.max_row, values_only=True):
        # Get the cell value from column B
        cell_value = row[2]

        if cell_value == "time":
            pass

        if row[0] == "Flight":
            pass

        if row[0] == ".":
            colC_list.append(".")

        else:
            new_date = f"{row[0]} {cell_value}"

            if new_date == "None None":
                pass
            else:
                # print(new_date)
                # Print the cell value
                df = pd.DataFrame(
                    columns=["datetime"],
                    data=pd.date_range(new_date, periods=1, freq="s"))

                df["minute_datetime"] = df["datetime"].dt.round("min")

                time = df["minute_datetime"].dt.strftime("%H:%M")

                contime = time.to_string()
                round_time = contime.split("   ")[1]
                colC_list.append(round_time.split(" ")[1])
    print(colC_list)

    for row, value in zip(worksheet.iter_rows(min_col=5, max_col=5, min_row=2), colC_list):
        row[0].value = value


def startZ_column_f():
    # Iterate through the rows of the worksheet
    for row in worksheet.iter_rows(min_row=2, max_col=4, max_row=worksheet.max_row, values_only=True):
        # Get the cell value from column B
        cell_value = row[3]

        if cell_value == "time":
            pass

        if row[0] == "Flight":
            pass

        if row[0] == ".":
            colF_list.append(".")

        else:
            new_date = f"{row[0]} {cell_value}"

            if new_date == "None None":
                pass

            else:
                # print(new_date)
                # Print the cell value
                df = pd.DataFrame(
                    columns=["datetime"],
                    data=pd.date_range(new_date, periods=1, freq="s"))

                time_string = new_date.split(" ")[1]
                time_format = "%H:%M"

                # convert the time string to a datetime object
                time = datetime.strptime(time_string, time_format)

                # subtract 2 hours from the time

                new_time = time - timedelta(hours=2)

                contime = str(new_time)
                round_time = contime.split(" ")[1]

                colF_list.append(round_time)
    print(colF_list)

    for row, value in zip(worksheet.iter_rows(min_col=6, max_col=6, min_row=2), colF_list):
        row[0].value = value


def endZ_column_g():
    # Iterate through the rows of the worksheet
    for row in worksheet.iter_rows(min_row=2, max_col=5, max_row=worksheet.max_row, values_only=True):
        # Get the cell value from column B
        cell_value = row[4]

        if cell_value == "time":
            pass

        if row[0] == "Flight":
            pass

        if row[0] == ".":
            colG_list.append(".")

        else:
            new_date = f"{row[0]} {cell_value}"

            if new_date == "None None":
                pass

            else:
                # print(new_date)
                # Print the cell value
                df = pd.DataFrame(
                    columns=["datetime"],
                    data=pd.date_range(new_date, periods=1, freq="s"))

                time_string = new_date.split(" ")[1]
                time_format = "%H:%M"

                # convert the time string to a datetime object
                time = datetime.strptime(time_string, time_format)

                # subtract 2 hours from the time

                new_time = time - timedelta(hours=2)

                contime = str(new_time)
                round_time = contime.split(" ")[1]

                colG_list.append(round_time)
    print(colG_list)

    for row, value in zip(worksheet.iter_rows(min_col=7, max_col=7, min_row=2), colG_list):
        row[0].value = value


def Duration():
    for row in worksheet.iter_rows(min_row=2):
        # get the time values in columns F and G
        time1 = row[5].value  # column F
        time2 = row[6].value  # column G

        # skip rows where either time value is blank
        if not time1 or not time2:
            continue

        # skip cells with a period
        if '.' in time1 or '.' in time2:
            continue

        # convert the time values to datetime.time objects
        time1 = datetime.strptime(time1, '%H:%M:%S').time()
        time2 = datetime.strptime(time2, '%H:%M:%S').time()

        # combine the time values with a dummy date
        datetime1 = datetime.combine(datetime.min, time1)
        datetime2 = datetime.combine(datetime.min, time2)

        # calculate the duration between the two times
        duration = datetime2 - datetime1

        # format the duration as HH:MM
        duration_str = (datetime.min + duration).strftime('%H:%M')

        # write the duration to column H
        row[7].value = duration_str


# Set the starting location


def Location():
    # Loop through each row
    last_location = worksheet['I1'].value

    for row in range(2, worksheet.max_row + 1):
        current_location = worksheet.cell(row=row, column=9).value

        # If the location changes, insert a new row and apply the light blue fill
        if current_location and current_location != last_location:
            worksheet.insert_rows(row)
            for col in range(1, worksheet.max_column + 1):
                worksheet.cell(row=row, column=col).value = ","
                worksheet.cell(row=row, column=col).fill = lightBlue_fill
            last_location = current_location


# def deleteExtraYellow():
#     # Set the black fill style
#   black_fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')

#   # Loop through the rows in reverse order
#   for row in range(worksheet.max_row, 1, -1):
#       fill_color = worksheet.cell(row=row, column=1).fill.start_color.index

#       # If the fill color is black, delete all rows before it
#       if fill_color == black_fill.start_color.index:
#           for delete_row in range(1, row):
#               worksheet.delete_rows(delete_row)
#           break

def changeCells():
    # Set the yellow and red fill styles
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')

    # Initialize the previous row's fill color
    prev_fill_color = None

    # Loop through the rows
    for row in range(1, worksheet.max_row + 1):
        fill_color = worksheet.cell(row=row, column=8).fill.start_color.index

        # If the current row's fill color is not yellow and the previous row's fill color is yellow, change the fill color of the cell in column H to red
        # if fill_color != yellow_fill.start_color.index and prev_fill_color == yellow_fill.start_color.index:
        #     worksheet.cell(row=row, column=7).fill = red_fill

        # # Update the previous row's fill color
        # prev_fill_color = fill_color


round_column_b()
round_column_c()
startZ_column_f()
endZ_column_g()
Duration()
# Location()
# deleteExtraYellow()


# Save the changes
workbook.save('existing_file.xlsx')

# load the workbook and select the active worksheet
wb = load_workbook('existing_file.xlsx')
ws = wb.active

# loop through each row in the worksheet and convert the time values in columns F and G
for row in ws.iter_rows(min_row=2, min_col=6, max_col=7):
    for cell in row:
        # check if the cell value is not empty or None
        if cell.value is not None and cell.value != '':
            try:
                # try to parse the cell value as a time string in HH:MM:SS format
                datetime_obj = datetime.strptime(cell.value, '%H:%M:%S')
                # format the datetime object as a string in HH:MM format
                cell.value = datetime_obj.strftime('%H:%M')
            except ValueError:
                # if the cell value is not in HH:MM:SS format, set the cell value to an empty string
                cell.value = ''
        else:
            # if the cell value is empty or None, set the cell value to an empty string
            cell.value = ''

# set the red fill style
red_fill = PatternFill(start_color='FFFF0001', end_color='FFFF0000', fill_type='solid')

# loop through each row in the worksheet and check the value of column D
for row in ws.iter_rows(min_row=2, min_col=6, max_col=6):
    for cell in row:
        # check if the cell value is "."
        if cell.value == ".":
            # change the fill color of the previous cell to red
            cell.offset(row=-1, column=0).fill = red_fill
            cell.offset(row=2, column=1).fill = red_fill

# loop through each row in the worksheet
for row in ws.iter_rows(min_row=2):
    # check if the color of the previous row is black
    if row[0].offset(row=-1).fill.start_color.index == 'FF000000':
        # check if the color of the current row is yellow
        if row[0].fill.start_color.index == 'FFFFFF00':
            # delete the row
            ws.delete_rows(row[0].row)

# define the range of columns to hide
start_col = 2  # column A
end_col = 5  # column C

# loop through each column in the range
for col in range(start_col, end_col + 1):
    col_letter = get_column_letter(col)
    ws.column_dimensions[col_letter].hidden = True

# define the range of columns to hide
start_col = 11  # column A
end_col = 17  # column C

# loop through each column in the range
for col in range(start_col, end_col + 1):
    col_letter = get_column_letter(col)
    ws.column_dimensions[col_letter].hidden = True

ws.column_dimensions['I'].auto_size = True


# save the updated workbook
wb.save('Processed.xlsx')


